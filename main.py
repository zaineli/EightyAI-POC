from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse, Response
from pydantic import BaseModel
from typing import List, Optional
import pytesseract
from PIL import Image
import pdf2image
import httpx
import uuid
import json
from datetime import datetime
from pathlib import Path
import logging
import asyncio
import shutil
import os
from dotenv import load_dotenv
import csv
import io
import pandas as pd
import numpy as np
import cv2
from pdfminer.high_level import extract_text
from io import BytesIO
# Add these if not already present
import pandas as pd
import shutil
import json
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import zipfile

from fastapi.responses import FileResponse  # add this import
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Border, Side

# Define base and ledger paths once
BASE_DIR = Path(__file__).resolve().parent
GLOBAL_LEDGER_DIR = BASE_DIR / "global_ledger"
GLOBAL_LEDGER_CSV = GLOBAL_LEDGER_DIR / "global_ledger.csv"
GLOBAL_LEDGER_XLSX = GLOBAL_LEDGER_DIR / "global_ledger.xlsx"


LEDGER_DIR = Path(__file__).resolve().parent / "ledger"
LEDGER_DIR.mkdir(exist_ok=True)
LEDGER_FILE = LEDGER_DIR / "ledger.xlsx"

# ...existing code...



# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="PDF OCR + LLM Processing API", version="1.0.0")

# Configure CORS for your Next.js frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],  # Your Next.js app URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# For Camelot table extraction - add this section
try:
    import camelot
    CAMELOT_AVAILABLE = True
    logger.info("✅ Camelot is available for table extraction")
except ImportError:
    CAMELOT_AVAILABLE = False
    logger.warning("❌ Camelot not available. Install with: pip install camelot-py[cv]")

# Also add the Excel functionality imports
try:
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("❌ OpenPyXL not available. Excel export will be limited.")


# Load environment variables
load_dotenv()
OPENROUTER_API_KEY = os.getenv("API_KEY")
OPENROUTER_MODEL = "deepseek/deepseek-chat-v3.1:free"  # DeepSeek model

# Configuration
BASE_DIR = Path("jobs")
MAX_FILE_SIZE = 10 * 1024 * 1024  

BASE_DIR.mkdir(exist_ok=True)

class ProcessRequest(BaseModel):
    system_prompt: str
    user_prompt: str = None

class JobResponse(BaseModel):
    job_id: str
    status: str
    message: str
    files_processed: int
    created_at: str

# Update the DEFAULT_SYSTEM_PROMPT to include cross-verification instructions
DEFAULT_SYSTEM_PROMPT = """
You are an intelligent document processing assistant specialized in extracting and comparing information from invoices and delivery notes with particular focus on table content and item matching.

IMPORTANT EXTRACTION RULES:
- NEVER calculate any values yourself. Only extract what is explicitly stated in the documents.
- Pay special attention to TABLE data which has been pre-extracted by our OCR system.
- For each document, extract the VAT amount and total amount EXACTLY as they appear on the receipt/invoice.
- Extract every line item from each document's itemized table precisely as written.
- Maintain exact numeric values and formatting as they appear in the original documents.

CRITICAL CROSS-VERIFICATION REQUIREMENTS:
- Match every item in invoices against corresponding items in delivery notes by comparing item names, descriptions, and quantities.
- Items in the invoice MUST match the items in the delivery note.
- Identify any items that appear in an invoice but not in the related delivery note.
- Identify any items that appear in a delivery note but not in the related invoice.
- Report exact item name discrepancies (different spellings, formats, descriptions).
- Flag quantity mismatches between the same item in different documents.
- Mention ALL anomalies separately in the final output.
- Always return ALL items that are detected as common between both documents AND those that are mismatched.

REQUIRED OUTPUT FORMAT:
At the end of your response, you MUST provide the extracted data in CSV format within clearly marked sections:

====CSV_DATA_START====
INVOICE_DATA:
Invoice date,Invoice ID,Customer name,Invoice amount (No VAT),VAT,Total Amount
2025-01-15,INV-001,ABC Company,1000.00,200.00,1200.00

DELIVERY_NOTE_DATA:
Delivery note date,Delivery note number,Invoice number,Invoice date,Customer name
2025-01-16,DN-001,INV-001,2025-01-15,ABC Company

ANOMALIES:
Anomaly Type
Item missing in delivery note: Widget A
Quantity mismatch for Item B: Invoice=5, Delivery=3
====CSV_DATA_END====

For invoices, extract EXACTLY these fields:
1. Invoice date
2. Invoice ID/Number
3. Customer name
4. Invoice amount excluding VAT (numeric only)
5. VAT amount (numeric only)
6. Total amount (numeric only)
7. All line items with their exact names, quantities, descriptions, unit prices and totals

For delivery notes, extract EXACTLY these fields:
1. Delivery note date
2. Delivery note number
3. Associated invoice number (if present)
4. Associated invoice date (if present)
5. Customer name
6. All items listed with their exact names, quantities and descriptions

For readability and CSV compatibility, format your response in a PROPERLY FORMATTED TEXT structure (not JSON). Use the following structure:

------------------------------------------------------------
DOCUMENT TYPE: INVOICE
Invoice Date: YYYY-MM-DD
Invoice Number: ABC123
Customer Name: Company Name Ltd.
Amount Excluding VAT: 1000.00
VAT Amount: 200.00
Total Amount: 1200.00
Items:
- Item 1 | Quantity: 5 | Unit Price: 100.00
- Item 2 | Quantity: 2 | Unit Price: 250.00

DOCUMENT TYPE: DELIVERY NOTE
Delivery Note Date: YYYY-MM-DD
Delivery Note Number: DN123
Associated Invoice Number: INV456
Associated Invoice Date: YYYY-MM-DD
Customer Name: Company Name Ltd.
Items:
- Item 1 | Quantity: 5
- Item 2 | Quantity: 2

ITEM CROSS-VERIFICATION
Common Items (in both documents):
- Item 1 | Invoice Qty: 5 | Delivery Qty: 5 | Status: Match
- Item 2 | Invoice Qty: 2 | Delivery Qty: 2 | Status: Match
- Item 5 | Invoice Qty: 10 | Delivery Qty: 8 | Status: Quantity Mismatch
- Widget A (Invoice) vs Widget-A (Delivery) | Status: Name Format Different

Missing in Invoice:
- Item 3 | Delivery Qty: 1

Missing in Delivery:
- Item 4 | Invoice Qty: 3

ANOMALIES (List All Separately):
- Item 3 present in delivery note but missing in invoice
- Item 4 present in invoice but missing in delivery note
- Name mismatch between "Widget A" (invoice) and "Widget-A" (delivery)
- Quantity mismatch for Item 5 (Invoice: 10, Delivery: 8)
------------------------------------------------------------
"""
INTERNAL_COLUMNS = [
    "anomaly_type",
    "invoice_date",
    "invoice_id",
    "customer_name",
    "invoice_amount_no_vat",
    "vat",
    "total_amount",
    "delivery_note_date",
    "delivery_note_number",
    "delivery_invoice_number",
    "delivery_invoice_date",
    "delivery_customer_name",
]

DISPLAY_HEADERS = [
    "Anomaly Type",
    "Invoice date",
    "Invoice ID",
    "Customer name",
    "Invoice amount (No VAT)",
    "VAT",
    "Total Amount",
    "Delivery note date",
    "Delivery note number",
    "Invoice number",
    "Invoice date",
    "Customer name",
]

def _read_ledger_df(csv_path: Path) -> pd.DataFrame:
    """
    Read the ledger CSV safely:
    - Only take the first 12 columns (trim any historical extras)
    - Normalize to INTERNAL_COLUMNS
    """
    if not csv_path.exists():
        return pd.DataFrame(columns=INTERNAL_COLUMNS)

    try:
        # Read first 12 columns by position (handles duplicate headers)
        df = pd.read_csv(csv_path, header=0, usecols=list(range(len(INTERNAL_COLUMNS))), engine="python")
    except Exception:
        # Fallback: raw read then trim
        df = pd.read_csv(csv_path, header=None, engine="python")
        df = df.iloc[:, :len(INTERNAL_COLUMNS)]

    # Ensure we have exactly 12 cols and correct names
    if df.shape[1] > len(INTERNAL_COLUMNS):
        df = df.iloc[:, :len(INTERNAL_COLUMNS)]
    elif df.shape[1] < len(INTERNAL_COLUMNS):
        # pad missing columns
        for _ in range(len(INTERNAL_COLUMNS) - df.shape[1]):
            df[df.shape[1]] = ""

    df.columns = INTERNAL_COLUMNS
    return df

def _write_ledger_df(df: pd.DataFrame, csv_path: Path) -> None:
    """
    Write with internal columns but display headers so CSV/Excel show your exact headings.
    """
    out = df[INTERNAL_COLUMNS].copy()
    out.to_csv(csv_path, index=False, header=DISPLAY_HEADERS)

def get_next_job_id() -> str:
    """Generate the next job ID based on existing jobs"""
    existing_jobs = [d for d in BASE_DIR.iterdir() if d.is_dir() and d.name.startswith("job-")]
    if not existing_jobs:
        return "job-1"
    
    job_numbers = []
    for job_dir in existing_jobs:
        try:
            job_num = int(job_dir.name.split("-")[1])
            job_numbers.append(job_num)
        except (IndexError, ValueError):
            continue
    
    next_num = max(job_numbers) + 1 if job_numbers else 1
    return f"job-{next_num}"

def create_job_structure(job_id: str) -> Path:
    """Create job directory structure"""
    job_dir = BASE_DIR / job_id
    job_dir.mkdir(exist_ok=True)
    
    # Create subdirectories
    (job_dir / "uploads").mkdir(exist_ok=True)
    (job_dir / "ocr_results").mkdir(exist_ok=True)
    (job_dir / "processed_results").mkdir(exist_ok=True)
    
    return job_dir

@app.get("/")
async def root():
    return {"message": "PDF OCR + LLM Processing API is running"}

@app.post("/upload-multiple-pdfs", response_model=JobResponse)
async def upload_multiple_pdfs(
    files: List[UploadFile] = File(...),
    system_prompt: str = Form(default=DEFAULT_SYSTEM_PROMPT),
    user_prompt: str = Form(default="Please analyze these documents collectively and provide a comprehensive summary with cross-document insights.")
):
    """
    Upload multiple PDF files, process them with OCR, and analyze collectively with Llama
    """
    try:
        # Validate files
        pdf_files = []
        for file in files:
            if file.content_type != "application/pdf":
                raise HTTPException(
                    status_code=400, 
                    detail=f"File {file.filename} is not a PDF. Only PDF files are allowed"
                )
            
            # Read file content
            file_content = await file.read()
            
            # Validate file size
            if len(file_content) > MAX_FILE_SIZE:
                raise HTTPException(
                    status_code=400,
                    detail=f"File {file.filename} exceeds {MAX_FILE_SIZE / (1024*1024)}MB limit"
                )
            
            pdf_files.append({
                "filename": file.filename,
                "content": file_content
            })
        
        # Generate job ID and create structure
        job_id = get_next_job_id()
        job_dir = create_job_structure(job_id)
        
        logger.info(f"Created job {job_id} for {len(pdf_files)} files")
        
        # Save job metadata
        job_metadata = {
            "job_id": job_id,
            "created_at": datetime.now().isoformat(),
            "total_files": len(pdf_files),
            "system_prompt": system_prompt,
            "user_prompt": user_prompt,
            "status": "processing",
            "files": [{"filename": f["filename"], "status": "pending"} for f in pdf_files]
        }
        
        with open(job_dir / "job_metadata.json", "w", encoding="utf-8") as f:
            json.dump(job_metadata, f, indent=2, ensure_ascii=False)
        
        # Process files asynchronously
        asyncio.create_task(process_job_files(job_id, job_dir, pdf_files, system_prompt, user_prompt))
        
        return JobResponse(
            job_id=job_id,
            status="processing",
            message=f"Job created successfully. Processing {len(pdf_files)} files.",
            files_processed=0,
            created_at=job_metadata["created_at"]
        )
        
    except Exception as e:
        logger.error(f"Error creating job: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error creating job: {str(e)}")

def initialize_global_ledger():
    """Initialize the global structured ledger files if they don't exist"""
    global_ledger_dir = Path("global_ledger")
    global_ledger_dir.mkdir(exist_ok=True)

    csv_path = global_ledger_dir / "global_ledger.csv"
    xlsx_path = global_ledger_dir / "global_ledger.xlsx"

    if not csv_path.exists():
        df = pd.DataFrame(columns=INTERNAL_COLUMNS)
        _write_ledger_df(df, csv_path)
        logger.info(f"Created global ledger CSV at {csv_path}")
    else:
        # Repair any legacy CSVs with extra columns and rewrite headers
        df = _read_ledger_df(csv_path)
        _write_ledger_df(df, csv_path)

    if not xlsx_path.exists() and OPENPYXL_AVAILABLE:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Global Ledger"
            for col_idx, header in enumerate(DISPLAY_HEADERS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                try:
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center")
                except Exception:
                    pass
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            wb.save(xlsx_path)
            logger.info(f"Created global ledger Excel at {xlsx_path}")
        except Exception as e:
            logger.error(f"Failed to create Excel global ledger: {str(e)}")


def update_global_ledger(extracted_data, job_id):
    """Update the global ledger with data from a specific job, placing all data in a single row"""
    global_ledger_dir = Path("global_ledger")
    csv_path = global_ledger_dir / "global_ledger.csv"
    xlsx_path = global_ledger_dir / "global_ledger.xlsx"

    initialize_global_ledger()

    try:
        # Always read normalized 12-column schema
        df = _read_ledger_df(csv_path)

        new_rows = []

        invoice_rows = extracted_data.get("csv_data", {}).get("invoice_rows", [])
        delivery_rows = extracted_data.get("csv_data", {}).get("delivery_note_rows", [])
        anomaly_rows = extracted_data.get("csv_data", {}).get("anomaly_rows", [])

        delivery_by_invoice = {}
        for delivery_row in delivery_rows:
            parts = [p.strip() for p in delivery_row.split(",")]
            if len(parts) >= 5 and parts[2]:
                delivery_by_invoice[parts[2]] = parts  # [dn_date, dn_no, inv_no, inv_date, cust_name]

        invoices_by_id = {}
        for invoice_row in invoice_rows:
            parts = [p.strip() for p in invoice_row.split(",")]
            if len(parts) >= 6:
                invoices_by_id[parts[1]] = parts  # [inv_date, inv_id, cust_name, amount_no_vat, vat, total]

        invoice_ids = set(invoices_by_id.keys()) | set(delivery_by_invoice.keys())

        for inv_id in invoice_ids:
            row = {k: "" for k in INTERNAL_COLUMNS}

            if inv_id in invoices_by_id:
                inv = invoices_by_id[inv_id]
                row["invoice_date"] = inv[0]
                row["invoice_id"] = inv_id
                row["customer_name"] = inv[2]
                row["invoice_amount_no_vat"] = inv[3]
                row["vat"] = inv[4]
                row["total_amount"] = inv[5]

            if inv_id in delivery_by_invoice:
                dn = delivery_by_invoice[inv_id]
                row["delivery_note_date"] = dn[0]
                row["delivery_note_number"] = dn[1]
                row["delivery_invoice_number"] = dn[2]
                row["delivery_invoice_date"] = dn[3]
                row["delivery_customer_name"] = dn[4]

            related = [a for a in anomaly_rows if inv_id and inv_id in a]
            if related:
                row["anomaly_type"] = "; ".join(related)

            new_rows.append(row)

        matched_ids = {r["invoice_id"] for r in new_rows if r.get("invoice_id")}
        orphaned = []
        for a in anomaly_rows:
            if not any(inv_id and inv_id in a for inv_id in matched_ids):
                orphaned.append(a)
        if orphaned:
            new_rows.append({k: "" for k in INTERNAL_COLUMNS} | {"anomaly_type": "; ".join(orphaned)})

        if new_rows:
            new_df = pd.DataFrame(new_rows, columns=INTERNAL_COLUMNS)
            result_df = pd.concat([df, new_df], ignore_index=True)

            # Write CSV with your exact headers
            _write_ledger_df(result_df, csv_path)
            logger.info(f"Updated global ledger CSV with {len(new_rows)} rows from job {job_id}")

            if xlsx_path.exists() and OPENPYXL_AVAILABLE:
                try:
                    wb = openpyxl.load_workbook(xlsx_path)
                    ws = wb.active
                    next_row = ws.max_row + 1

                    for r in new_rows:
                        for col_idx, key in enumerate(INTERNAL_COLUMNS, start=1):
                            val = r.get(key, "")
                            ws.cell(row=next_row, column=col_idx, value=val)

                            if key == "anomaly_type" and val:
                                try:
                                    ws.cell(row=next_row, column=col_idx).fill = openpyxl.styles.PatternFill(
                                        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                                    )
                                    ws.cell(row=next_row, column=col_idx).font = openpyxl.styles.Font(color="9C6500")
                                except Exception as style_error:
                                    logger.warning(f"Could not apply style: {style_error}")

                        try:
                            border = openpyxl.styles.Border(
                                left=openpyxl.styles.Side(style="thin"),
                                right=openpyxl.styles.Side(style="thin"),
                                top=openpyxl.styles.Side(style="thin"),
                                bottom=openpyxl.styles.Side(style="thin"),
                            )
                            for col_idx in range(1, len(INTERNAL_COLUMNS) + 1):
                                ws.cell(row=next_row, column=col_idx).border = border
                        except Exception:
                            pass

                        next_row += 1

                    wb.save(xlsx_path)
                    logger.info(f"Updated global ledger Excel with {len(new_rows)} rows from job {job_id}")
                except Exception as e:
                    logger.error(f"Failed to update Excel global ledger: {str(e)}")

            return result_df
        else:
            logger.warning(f"No data to add to global ledger from job {job_id}")
            return df

    except Exception as e:
        logger.error(f"Error updating global ledger: {str(e)}")
        raise
# Replace the existing startup_tasks function with this one
@app.on_event("startup")
async def startup_tasks():
    """Tasks to run on server startup"""
    # Initialize the global ledger
    initialize_global_ledger()
    
    # Check dependencies (removing the missing function call)
    try:
        # Test Tesseract
        pytesseract.get_tesseract_version()
        logger.info("✅ Tesseract OCR is available")
        
        # Test Camelot for table extraction
        if CAMELOT_AVAILABLE:
            logger.info("✅ Camelot table extraction is available")
        else:
            logger.warning("❌ Camelot is not available. Table extraction will be limited.")
            logger.info("💡 To install Camelot: pip install camelot-py[cv]")
            
        # Test PDF to image conversion (indirectly tests poppler)
        try:
            import pdf2image
            logger.info("✅ PDF2Image is available")
        except ImportError:
            logger.warning("❌ PDF2Image not available. Install poppler.")
            
        # Test openpyxl
        if OPENPYXL_AVAILABLE:
            logger.info("✅ OpenPyXL is available for Excel export")
        else:
            logger.warning("❌ OpenPyXL not available. Excel export will use basic formatting.")
            logger.info("💡 To install OpenPyXL: pip install openpyxl")
            
    except Exception as e:
        logger.error(f"❌ Dependency check failed: {e}")
        logger.error("Please install tesseract (apt install tesseract-ocr on Debian/Ubuntu)")

def extract_csv_data_from_response(response_text: str) -> dict:
    """Extract structured CSV data from the LLM response"""
    result = {
        "csv_data": {
            "invoice_rows": [],
            "delivery_note_rows": [],
            "anomaly_rows": []
        }
    }
    
    try:
        # Check if response contains CSV data section
        if "====CSV_DATA_START====" not in response_text or "====CSV_DATA_END====" not in response_text:
            logger.warning("No CSV data markers found in LLM response")
            return result
        
        # Extract the CSV data section
        csv_section = response_text.split("====CSV_DATA_START====")[1].split("====CSV_DATA_END====")[0].strip()
        
        # Process each section
        current_section = None
        header_skipped = False
        
        for line in csv_section.split("\n"):
            line = line.strip()
            if not line:
                continue
                
            # Identify sections
            if line.startswith("INVOICE_DATA:"):
                current_section = "invoice"
                header_skipped = False
                continue
            elif line.startswith("DELIVERY_NOTE_DATA:"):
                current_section = "delivery_note"
                header_skipped = False
                continue
            elif line.startswith("ANOMALIES:"):
                current_section = "anomaly"
                header_skipped = False
                continue
            
            # Skip header rows
            if not header_skipped:
                header_skipped = True
                continue
            
            # Store data in appropriate section
            if current_section == "invoice":
                result["csv_data"]["invoice_rows"].append(line)
            elif current_section == "delivery_note":
                result["csv_data"]["delivery_note_rows"].append(line)
            elif current_section == "anomaly":
                result["csv_data"]["anomaly_rows"].append(line)
    
        logger.info(f"Extracted {len(result['csv_data']['invoice_rows'])} invoice rows, "
                   f"{len(result['csv_data']['delivery_note_rows'])} delivery note rows, and "
                   f"{len(result['csv_data']['anomaly_rows'])} anomaly rows")
                   
    except Exception as e:
        logger.error(f"Error extracting CSV data from response: {str(e)}")
        
    return result

# 2. Fix the process_job_files function to better handle OCR failures
async def process_job_files(job_id: str, job_dir: Path, pdf_files: List, system_prompt: str, user_prompt: str):
    """Process all files in a job asynchronously"""
    try:
        logger.info(f"Starting processing for job {job_id}")
        
        # Update job status
        update_job_status(job_dir, "processing")
        
        all_ocr_results = []
        processed_files = []
        failed_files = []
        
        # Process each PDF file
        for i, pdf_file in enumerate(pdf_files, 1):
            try:
                logger.info(f"Processing file {i}/{len(pdf_files)}: {pdf_file['filename']}")
                
                # Generate unique filename and save
                file_extension = Path(pdf_file["filename"]).suffix
                unique_filename = f"file_{i}_{uuid.uuid4()}{file_extension}"
                file_path = job_dir / "uploads" / unique_filename
                
                with open(file_path, "wb") as buffer:
                    buffer.write(pdf_file["content"])
                
                logger.info(f"Saved file to {file_path}")
                
                # Process with enhanced OCR
                try:
                    logger.info(f"Starting enhanced OCR for {file_path}")
                    ocr_result = await process_pdf_with_enhanced_ocr(
                        file_path, 
                        f"{job_id}_file_{i}", 
                        job_dir / "ocr_results",
                        original_filename=pdf_file["filename"]
                    )
                    
                    # Verify OCR result has content
                    if not ocr_result or not ocr_result.get("full_text", "").strip():
                        logger.warning(f"OCR produced empty result for {pdf_file['filename']}")
                        ocr_result["full_text"] = f"[OCR EXTRACTION FAILED FOR {pdf_file['filename']}]"
                    
                    all_ocr_results.append(ocr_result)
                    processed_files.append({
                        "original_filename": pdf_file["filename"],
                        "stored_filename": unique_filename,
                        "file_number": i,
                        "ocr_result": ocr_result,
                        "status": "success",
                        "tables_extracted": len(ocr_result.get("tables", [])) 
                    })
                    
                    logger.info(f"Completed enhanced OCR for file {i}/{len(pdf_files)} with {len(ocr_result.get('tables', []))} tables extracted")
                except Exception as ocr_error:
                    logger.error(f"OCR processing failed for {pdf_file['filename']}: {str(ocr_error)}")
                    failed_files.append({
                        "original_filename": pdf_file["filename"],
                        "stored_filename": unique_filename,
                        "file_number": i,
                        "error": str(ocr_error),
                        "status": "failed"
                    })
                    continue
                
            except Exception as e:
                logger.error(f"Error processing file {pdf_file['filename']}: {str(e)}")
                failed_files.append({
                    "original_filename": pdf_file["filename"],
                    "file_number": i,
                    "error": str(e),
                    "status": "failed"
                })
                continue
        
        # Check if we have any successful OCR results
        if not processed_files:
            error_msg = "All files failed OCR processing"
            logger.error(error_msg)
            update_job_status(job_dir, "failed", error_msg)
            return
        
        # Combine all OCR results for collective analysis
        combined_context = prepare_combined_context(processed_files)
        
        # Save the combined context for debugging
        with open(job_dir / "combined_context.txt", "w", encoding="utf-8") as f:
            f.write(combined_context)
        
        # Process with OpenRouter
        openrouter_response = await process_with_openrouter(
            combined_context,
            system_prompt,
            user_prompt,
            job_id,
            job_dir / "processed_results"
        )
        
        # Handle case where OpenRouter might fail
        if isinstance(openrouter_response, Exception):
            logger.error(f"OpenRouter processing failed: {str(openrouter_response)}")
            openrouter_response = {
                "error": str(openrouter_response),
                "response": f"Error processing with OpenRouter: {str(openrouter_response)}"
            }
        
        # Extract CSV data from the response
        extracted_data = extract_csv_data_from_response(openrouter_response.get('response', ''))

        # Update the global ledger
        update_global_ledger(extracted_data, job_id)

        # If you also want to update the specific ledger file, add this call:
        update_ledger_with_extracted_data(extracted_data)
        
        # Save extracted CSV data for debugging
        with open(job_dir / "extracted_csv_data.json", "w", encoding="utf-8") as f:
            json.dump(extracted_data, f, indent=2, ensure_ascii=False)
        
        # Use standardized response structure
        llm_response = {
            "openrouter": openrouter_response,
            "model_used": OPENROUTER_MODEL,
            "response": openrouter_response.get('response', 'Processing failed'),
            "total_tokens": openrouter_response.get("total_tokens", 0),
            "context_length": len(combined_context),
            "response_length": len(openrouter_response.get("response", ""))
        }
        
        # Update final job status with standardized structure
        job_summary = {
            "job_id": job_id,
            "status": "completed",
            "completed_at": datetime.now().isoformat(),
            "total_files": len(pdf_files),
            "successfully_processed": len(processed_files),
            "failed_files": len(pdf_files) - len(processed_files),
            "processed_files": processed_files,
            "llm_analysis": llm_response,
            "extracted_csv_data": extracted_data
        }
        
        # Save results in both files for consistency
        with open(job_dir / "job_summary.json", "w", encoding="utf-8") as f:
            json.dump(job_summary, f, indent=2, ensure_ascii=False)
            
        with open(job_dir / "job_results.json", "w", encoding="utf-8") as f:
            json.dump(job_summary, f, indent=2, ensure_ascii=False)
        
        update_job_status(job_dir, "completed")
        
        logger.info(f"Job {job_id} completed successfully")
        
        # Add at the end of process_job_files function
        # Update the global ledger with the extracted data
        update_global_ledger(extracted_data, job_id)
        
    except Exception as e:
        logger.error(f"Error processing job {job_id}: {str(e)}")
        update_job_status(job_dir, "failed", str(e))

def update_job_status(job_dir: Path, status: str, error_message: str = None):
    """Update job status in metadata"""
    metadata_file = job_dir / "job_metadata.json"
    try:
        with open(metadata_file, "r", encoding="utf-8") as f:
            metadata = json.load(f)
        
        metadata["status"] = status
        metadata["last_updated"] = datetime.now().isoformat()
        
        if error_message:
            metadata["error"] = error_message
        
        with open(metadata_file, "w", encoding="utf-8") as f:
            json.dump(metadata, f, indent=2, ensure_ascii=False)
    except Exception as e:
        logger.error(f"Failed to update job status: {str(e)}")

def prepare_combined_context(processed_files: List[dict]) -> str:
    """Prepare combined context from all processed files with emphasis on tables"""
    combined_parts = []
    
    for file in processed_files:
        ocr_result = file.get("ocr_result", {})
        original_filename = file.get("original_filename", "Unknown file")
        
        combined_parts.append(f"## DOCUMENT: {original_filename}")
        
        # Add table data in a structured format
        tables = ocr_result.get("tables", [])
        if tables:
            combined_parts.append(f"\n### TABLES ({len(tables)} found):")
            for table in tables:
                combined_parts.append(f"\n#### TABLE {table['table_number']}:")
                
                # Convert table data to formatted text
                rows = table.get('data', [])
                if rows:
                    table_text = ""
                    for i, row in enumerate(rows):
                        row_text = " | ".join([str(cell).strip() for cell in row])
                        if i == 0 and i < len(rows) - 1:  # Likely header
                            row_text = f"{row_text}\n{'-' * len(row_text)}"
                        table_text += f"{row_text}\n"
                    combined_parts.append(table_text)
        
        # Add the full text
        combined_parts.append("\n### FULL TEXT:")
        combined_parts.append(ocr_result.get("full_text", "[No text extracted]"))
        
        combined_parts.append("\n" + "=" * 80 + "\n")
    
    return "\n\n".join(combined_parts)

def clean_ocr_text(text: str) -> str:
    """Clean and normalize OCR text for better processing"""
    if not text:
        return ""
    
    # Remove excessive whitespace but preserve structure
    lines = text.split('\n')
    cleaned_lines = []
    
    for line in lines:
        # Remove leading/trailing whitespace
        cleaned_line = line.strip()
        # Replace multiple spaces with single space
        cleaned_line = ' '.join(cleaned_line.split())
        if cleaned_line:  # Only add non-empty lines
            cleaned_lines.append(cleaned_line)
    
    return '\n'.join(cleaned_lines)

async def process_pdf_with_ocr(pdf_path: Path, file_id: str, ocr_results_dir: Path, original_filename: str = None):
    """Convert PDF to images and apply OCR using Tesseract"""
    try:
        # Convert PDF to images
        logger.info(f"Converting PDF to images: {pdf_path}")
        pages = pdf2image.convert_from_path(
            pdf_path,
            dpi=300, 
            fmt='PNG'
        )
        
        ocr_results = {
            "file_id": file_id,
            "original_filename": original_filename or pdf_path.name,
            "stored_filename": pdf_path.name,
            "total_pages": len(pages),
            "pages": [],
            "tables": [],
            "full_text": "",
            "processing_timestamp": datetime.now().isoformat()
        }
        
        full_text_parts = []
        
        # Process each page
        for page_num, page_image in enumerate(pages, 1):
            logger.info(f"Processing page {page_num}/{len(pages)} for {original_filename}")
            
            page_text = pytesseract.image_to_string(
                page_image,
                config='--oem 3 --psm 6' 
            )
            
            # Get detailed OCR data
            page_data = pytesseract.image_to_data(
                page_image,
                output_type=pytesseract.Output.DICT,
                config='--oem 3 --psm 6'
            )
            
            # Filter out empty text and low confidence results
            filtered_words = []
            for i in range(len(page_data['text'])):
                if int(page_data['conf'][i]) > 30 and page_data['text'][i].strip():
                    filtered_words.append({
                        'text': page_data['text'][i],
                        'confidence': int(page_data['conf'][i]),
                        'bbox': {
                            'x': page_data['left'][i],
                            'y': page_data['top'][i],
                            'width': page_data['width'][i],
                            'height': page_data['height'][i]
                        }
                    })
            
            page_result = {
                "page_number": page_num,
                "text": page_text.strip(),
                "word_count": len(page_text.split()),
                "words_with_positions": filtered_words,
                "average_confidence": sum(word['confidence'] for word in filtered_words) / len(filtered_words) if filtered_words else 0
            }
            
            ocr_results["pages"].append(page_result)
            full_text_parts.append(page_text.strip())
        
        # Combine all text
        ocr_results["full_text"] = "\n\n".join(full_text_parts)
        ocr_results["total_words"] = len(ocr_results["full_text"].split())
        
        # Save OCR results
        results_file = ocr_results_dir / f"{file_id}_ocr_results.json"
        with open(results_file, "w", encoding="utf-8") as f:
            json.dump(ocr_results, f, indent=2, ensure_ascii=False)
        
        return ocr_results
        
    except Exception as e:
        logger.error(f"Error in OCR processing for {original_filename}: {str(e)}")
        raise Exception(f"OCR processing failed for {original_filename}: {str(e)}")

async def process_pdf_with_enhanced_ocr(pdf_path: Path, file_id: str, ocr_results_dir: Path, original_filename: str = None):
    """Process PDF with enhanced OCR including table structure extraction"""
    try:
        # Check if file exists
        if not pdf_path.exists():
            raise Exception(f"PDF file does not exist at {pdf_path}")
            
        # Check file size
        file_size = pdf_path.stat().st_size
        logger.info(f"PDF file size: {file_size} bytes")
        
        # Create results directory if it doesn't exist
        ocr_results_dir.mkdir(exist_ok=True, parents=True)
        
        # 1. First extract basic text with pdfminer for overall context
        logger.info(f"Extracting basic text from {pdf_path}")
        basic_text = extract_text(str(pdf_path))
        
        # 2. Convert PDF to images for Tesseract OCR
        logger.info(f"Converting PDF to images for Tesseract: {pdf_path}")
        try:
            pages = pdf2image.convert_from_path(
                str(pdf_path),
                dpi=300, 
                fmt='PNG'
            )
            logger.info(f"Successfully converted {len(pages)} pages to images")
        except Exception as e:
            logger.error(f"PDF to image conversion failed: {str(e)}")
            raise Exception(f"PDF to image conversion failed: {str(e)}")
        
        if not pages:
            raise Exception("PDF conversion produced no images")
        
        # 3. Try to extract tables with Camelot
        logger.info(f"Extracting tables with Camelot from {pdf_path}")
        tables_data = []
        tables_html = []
        
        if CAMELOT_AVAILABLE:
            try:
                # Try lattice mode first (for tables with borders)
                tables = camelot.read_pdf(str(pdf_path), flavor='lattice', pages='all')
                
                # If no tables found, try stream mode (for tables without borders)
                if len(tables) == 0:
                    tables = camelot.read_pdf(str(pdf_path), flavor='stream', pages='all')
                    
                logger.info(f"Camelot found {len(tables)} tables")
                
                for i, table in enumerate(tables):
                    # Check table quality
                    accuracy = table.parsing_report.get('accuracy', 0)
                    if accuracy < 50:  # Filter low-quality tables
                        continue
                        
                    # Convert to dictionary representation
                    df = table.df
                    tables_data.append({
                        'table_number': i+1,
                        'rows': df.shape[0],
                        'columns': df.shape[1],
                        'data': df.values.tolist(),
                        'headers': df.iloc[0].values.tolist() if not df.empty else [],
                        'accuracy': accuracy
                    })
                    
                    # Also save as HTML for visualization
                    html = df.to_html(index=False)
                    tables_html.append(html)
                    
            except Exception as e:
                logger.warning(f"Camelot table extraction failed: {str(e)}")
                # Continue with regular OCR even if table extraction fails
        else:
            logger.warning("Camelot not available, skipping table extraction")

        # 4. Process each page with Tesseract for detailed text
        ocr_results = {
            "file_id": file_id,
            "original_filename": original_filename or pdf_path.name,
            "stored_filename": pdf_path.name,
            "total_pages": len(pages),
            "pages": [],
            "tables": tables_data,
            "full_text": basic_text,  # Use pdfminer text as base
            "processing_timestamp": datetime.now().isoformat()
        }
        
        # Save table HTML representations for debugging/visualization
        if tables_html:
            tables_html_path = ocr_results_dir / f"{file_id}_tables.html"
            with open(tables_html_path, "w", encoding="utf-8") as f:
                f.write("<html><body>")
                f.write(f"<h1>Tables from {original_filename}</h1>")
                for i, html in enumerate(tables_html):
                    f.write(f"<h2>Table {i+1}</h2>")
                    f.write(html)
                f.write("</body></html>")
            
            logger.info(f"Saved table visualization to {tables_html_path}")
        
        # Process each page with Tesseract for any text not in tables
        for page_num, page_image in enumerate(pages, 1):
            logger.info(f"Processing OCR for page {page_num}/{len(pages)} for {original_filename}")
            
            try:
                # Save the image for debugging
                img_debug_path = ocr_results_dir / f"{file_id}_page_{page_num}.png"
                page_image.save(img_debug_path)
                
                # Enhance image for better OCR
                img = np.array(page_image)
                gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                # Apply adaptive thresholding to better isolate text
                binary = cv2.adaptiveThreshold(
                    gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                    cv2.THRESH_BINARY, 11, 2
                )
                
                # Process with Tesseract with optimized settings for invoices
                page_text = pytesseract.image_to_string(
                    binary,
                    config='--oem 3 --psm 6 -l eng'  # Optimized for mixed content
                )
                
                # Also get word-level data with positions
                word_data = pytesseract.image_to_data(
                    binary, 
                    output_type=pytesseract.Output.DICT,
                    config='--oem 3 --psm 6 -l eng'
                )
                
                # Filter out low-confidence words
                conf_threshold = 60  # confidence threshold
                words_with_positions = []
                total_conf = 0
                word_count = 0
                
                for i in range(len(word_data['text'])):
                    if int(word_data['conf'][i]) > conf_threshold and word_data['text'][i].strip():
                        word_info = {
                            'text': word_data['text'][i],
                            'conf': int(word_data['conf'][i]),
                            'x': word_data['left'][i],
                            'y': word_data['top'][i],
                            'width': word_data['width'][i],
                            'height': word_data['height'][i]
                        }
                        words_with_positions.append(word_info)
                        total_conf += int(word_data['conf'][i])
                        word_count += 1

                # CORRECT syntax:
                avg_conf = total_conf / word_count if word_count > 0 else 0

                # Store page results
                ocr_results["pages"].append({
                    "page_number": page_num,
                    "text": page_text,
                    "word_count": word_count,
                    "words_with_positions": words_with_positions,
                    "average_confidence": avg_conf
                })
                
                logger.info(f"OCR extracted {word_count} words from page {page_num} with avg confidence {avg_conf:.1f}%")
                
            except Exception as e:
                logger.error(f"Error in OCR processing for page {page_num}: {str(e)}")
                ocr_results["pages"].append({
                    "page_number": page_num,
                    "text": f"[OCR ERROR: {str(e)}]",
                    "word_count": 0,
                    "words_with_positions": [],
                    "average_confidence": 0
                })
        
        # Merge table data and OCR text for the most comprehensive results
        all_text_parts = [ocr_results["full_text"]]
        
        # Extract text from tables to ensure nothing is missed
        for table in tables_data:
            table_text = "\nTABLE:\n"
            for row in table['data']:
                table_text += " | ".join([str(cell).strip() for cell in row]) + "\n"
            all_text_parts.append(table_text)
        
        # Add full text from OCR
        ocr_results["full_text"] = "\n\n".join(all_text_parts)
        
        # Calculate total words
        total_words = sum(page["word_count"] for page in ocr_results["pages"])
        ocr_results["total_words"] = total_words
        
        # Save OCR results
        results_file = ocr_results_dir / f"{file_id}_ocr_results.json"
        with open(results_file, "w", encoding="utf-8") as f:
            json.dump(ocr_results, f, indent=2, ensure_ascii=False)
        
        logger.info(f"OCR completed for {pdf_path}. Extracted {total_words} words and {len(tables_data)} tables.")
        return ocr_results
        
    except Exception as e:
        logger.error(f"Error in OCR processing for {pdf_path}: {str(e)}")
        raise Exception(f"OCR processing failed for {original_filename or pdf_path.name}: {str(e)}")

# async def process_with_llama_collective(combined_text: str, system_prompt: str, user_prompt: str, job_id: str, results_dir: Path):
#     """Process the combined extracted text with local Llama model"""
#     try:
#         logger.info(f"Processing job {job_id} with Llama model...")
        
#         # Ensure we don't exceed context limits
#         max_context_length = 120000  # Conservative limit for llama3.2:3b
#         if len(combined_text) > max_context_length:
#             logger.warning(f"Content length ({len(combined_text)}) exceeds limit, truncating...")
#             combined_text = combined_text[:max_context_length] + "\n\n[CONTENT TRUNCATED DUE TO LENGTH LIMITS]"
        
#         # Enhanced user prompt for better structured output
#         enhanced_user_prompt = f"""
# {user_prompt}

# Please analyze the provided multi-document content and extract structured information from each document.

# REQUIREMENTS:
# 1. Process each document individually and extract all relevant data
# 2. Present findings in a clear, well-organized format
# 3. Always specify which document (number and filename) contains each piece of information
# 4. Use consistent formatting for dates, numbers, and currency
# 5. Identify any relationships between documents
# 6. Provide a summary section at the end

# Multi-document content follows below:

# {combined_text}
# """
        
#         # Prepare request payload for Ollama
#         payload = {
#             "model": MODEL_NAME,
#             "messages": [
#                 {"role": "system", "content": system_prompt},
#                 {"role": "user", "content": enhanced_user_prompt}
#             ],
#             "stream": False,
#             "options": {
#                 "temperature": 0.3,  # Lower temperature for more consistent output
#                 "top_p": 0.9,
#                 "max_tokens": 8192,  # Increased for detailed analysis
#                 "stop": ["[END_ANALYSIS]"]
#             }
#         }
        
#         logger.info(f"Sending request to Ollama - Context length: {len(enhanced_user_prompt):,} chars")
        
#         # Make request to Ollama with longer timeout
#         async with httpx.AsyncClient(timeout=900.0) as client:  # 15 minutes timeout
#             response = await client.post(
#                 f"{OLLAMA_URL}/api/chat",
#                 json=payload
#             )
            
#             if response.status_code != 200:
#                 error_detail = f"Ollama API error: {response.status_code}"
#                 try:
#                     error_data = response.json()
#                     error_detail += f" - {error_data}"
#                 except:
#                     error_detail += f" - {response.text[:500]}"
#                 raise Exception(error_detail)
            
#             result = response.json()
            
#             if "message" not in result or "content" not in result["message"]:
#                 raise Exception(f"Unexpected Ollama response format: {result}")
        
#         # Format and clean the response
#         raw_response = result["message"]["content"]
#         formatted_response = format_llm_response(raw_response)
        
#         llm_response = {
#             "job_id": job_id,
#             "model_used": MODEL_NAME,
#             "system_prompt": system_prompt,
#             "user_prompt": user_prompt,
#             "response": formatted_response,
#             "raw_response": raw_response,  # Keep original for debugging
#             "processing_timestamp": datetime.now().isoformat(),
#             "total_tokens": result.get("eval_count", 0),
#             "prompt_tokens": result.get("prompt_eval_count", 0),
#             "context_length": len(combined_text),
#             "response_length": len(formatted_response)
#         }
        
#         # Save LLM results
#         results_file = results_dir / f"{job_id}_collective_analysis.json"
#         with open(results_file, "w", encoding="utf-8") as f:
#             json.dump(llm_response, f, indent=2, ensure_ascii=False)
        
#         logger.info(f"Llama collective processing completed for job {job_id} - Response length: {len(formatted_response):,} chars")
        
#         return llm_response
        
#     except Exception as e:
#         logger.error(f"Error in Llama collective processing: {str(e)}")
#         raise Exception(f"Llama collective processing failed: {str(e)}")

async def process_with_openrouter(combined_text: str, system_prompt: str, user_prompt: str, job_id: str, results_dir: Path):
    """Process the combined extracted text with OpenRouter API"""
    try:
        logger.info(f"Processing job {job_id} with OpenRouter model...")
        
        # Ensure we don't exceed context limits
        max_context_length = 100000  # Conservative limit for OpenRouter
        if len(combined_text) > max_context_length:
            logger.warning(f"Content length ({len(combined_text)}) exceeds limit, truncating...")
            combined_text = combined_text[:max_context_length] + "\n\n[CONTENT TRUNCATED DUE TO LENGTH LIMITS]"
        
        # Use the same enhanced user prompt for consistency
        enhanced_user_prompt = f"""
{user_prompt}

Please analyze the provided multi-document content and extract structured information from each document.

REQUIREMENTS:
1. Process each document individually and extract all relevant data
2. Present findings in a clear, well-organized format
3. Always specify which document (number and filename) contains each piece of information
4. Use consistent formatting for dates, numbers, and currency
5. Identify any relationships between documents
6. Provide a summary section at the end

Multi-document content follows below:

{combined_text}
"""
        
        # Prepare request payload for OpenRouter
        payload = {
            "model": OPENROUTER_MODEL,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": enhanced_user_prompt}
            ]
        }
        
        # Headers for OpenRouter
        headers = {
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
            "HTTP-Referer": "http://localhost:3000",  # Your frontend URL
            "X-Title": "Intelligent Document Processor"  # App name
        }
        
        logger.info(f"Sending request to OpenRouter - Context length: {len(enhanced_user_prompt):,} chars")
        
        # Make request to OpenRouter API
        async with httpx.AsyncClient(timeout=300.0) as client:  # 5 minutes timeout
            response = await client.post(
                "https://openrouter.ai/api/v1/chat/completions",
                json=payload,
                headers=headers
            )
            
            if response.status_code != 200:
                error_detail = f"OpenRouter API error: {response.status_code}"
                try:
                    error_data = response.json()
                    error_detail += f" - {error_data}"
                except:
                    error_detail += f" - {response.text[:500]}"
                raise Exception(error_detail)
            
            result = response.json()
            
            if "choices" not in result or not result["choices"] or "message" not in result["choices"][0]:
                raise Exception(f"Unexpected OpenRouter response format: {result}")
        
        # Format and clean the response
        raw_response = result["choices"][0]["message"]["content"]
        formatted_response = format_llm_response(raw_response)
        
        openrouter_response = {
            "job_id": job_id,
            "model_used": OPENROUTER_MODEL,
            "system_prompt": system_prompt,
            "user_prompt": user_prompt,
            "response": formatted_response,
            "raw_response": raw_response,  # Keep original for debugging
            "processing_timestamp": datetime.now().isoformat(),
            "total_tokens": result.get("usage", {}).get("total_tokens", 0),
            "prompt_tokens": result.get("usage", {}).get("prompt_tokens", 0),
            "context_length": len(combined_text),
            "response_length": len(formatted_response)
        }
        
        # Save OpenRouter results
        results_file = results_dir / f"{job_id}_openrouter_analysis.json"
        with open(results_file, "w", encoding="utf-8") as f:
            json.dump(openrouter_response, f, indent=2, ensure_ascii=False)
        
        logger.info(f"OpenRouter processing completed for job {job_id} - Response length: {len(formatted_response):,} chars")
        
        return openrouter_response
        
    except Exception as e:
        logger.error(f"Error in OpenRouter processing: {str(e)}")
        return {
            "job_id": job_id,
            "model_used": OPENROUTER_MODEL,
            "error": str(e),
            "response": f"Error processing with OpenRouter: {str(e)}",
            "processing_timestamp": datetime.now().isoformat()
        }

def format_llm_response(response: str) -> str:
    """Format the LLM response for better presentation"""
    if not response:
        return "No response received from the model."
    
    # Basic cleanup
    formatted = response.strip()
    
    # Ensure proper section spacing
    formatted = formatted.replace('\n##', '\n\n##')
    formatted = formatted.replace('\n###', '\n\n###')
    
    # Clean up any excessive newlines but preserve intentional spacing
    while '\n\n\n' in formatted:
        formatted = formatted.replace('\n\n\n', '\n\n')
    
    return formatted

@app.get("/job-status/{job_id}")
async def get_job_status(job_id: str):
    """Get current status of a job"""
    try:
        job_dir = BASE_DIR / job_id
        if not job_dir.exists():
            raise HTTPException(status_code=404, detail="Job not found")
        
        metadata_file = job_dir / "job_metadata.json"
        if metadata_file.exists():
            with open(metadata_file, "r", encoding="utf-8") as f:
                metadata = json.load(f)
            return JSONResponse(content=metadata)
        
        raise HTTPException(status_code=404, detail="Job metadata not found")
        
    except Exception as e:
        logger.error(f"Error retrieving job status: {str(e)}")
        raise HTTPException(status_code=500, detail="Error retrieving job status")

@app.get("/job-results/{job_id}")
async def get_job_results(job_id: str):
    """Get complete results of a completed job"""
    try:
        job_dir = BASE_DIR / job_id
        if not job_dir.exists():
            raise HTTPException(status_code=404, detail="Job not found")
        
        # First try job_results.json (used by ledger jobs)
        results_file = job_dir / "job_results.json"
        if results_file.exists():
            with open(results_file, "r", encoding="utf-8") as f:
                results = json.load(f)
            return JSONResponse(content=results)
        
        # Fallback to job_summary.json (used by regular jobs)
        summary_file = job_dir / "job_summary.json"
        if summary_file.exists():
            with open(summary_file, "r", encoding="utf-8") as f:
                summary = json.load(f)
            return JSONResponse(content=summary)
        
        raise HTTPException(status_code=404, detail="Job results not ready yet")
        
    except Exception as e:
        logger.error(f"Error retrieving job results: {str(e)}")
        raise HTTPException(status_code=500, detail="Error retrieving job results")

@app.get("/jobs")
async def list_all_jobs():
    """List all jobs with their current status"""
    try:
        jobs = []
        for job_dir in BASE_DIR.iterdir():
            if job_dir.is_dir() and job_dir.name.startswith("job-"):
                metadata_file = job_dir / "job_metadata.json"
                if metadata_file.exists():
                    with open(metadata_file, "r", encoding="utf-8") as f:
                        metadata = json.load(f)
                    jobs.append(metadata)
        
        # Sort by creation date (newest first)
        jobs.sort(key=lambda x: x.get("created_at", ""), reverse=True)
        
        return JSONResponse(content={"jobs": jobs})
        
    except Exception as e:
        logger.error(f"Error listing jobs: {str(e)}")
        raise HTTPException(status_code=500, detail="Error listing jobs")

@app.delete("/job/{job_id}")
async def delete_job(job_id: str):
    """Delete a job and all its associated files"""
    try:
        job_dir = BASE_DIR / job_id
        if not job_dir.exists():
            raise HTTPException(status_code=404, detail="Job not found")
        
        shutil.rmtree(job_dir)
        
        return JSONResponse(content={"message": f"Job {job_id} deleted successfully"})
        
    except Exception as e:
        logger.error(f"Error deleting job: {str(e)}")
        raise HTTPException(status_code=500, detail="Error deleting job")

# @app.get("/health/ollama")
# async def check_ollama_health():
#     """Check if Ollama is running and the model is available"""
#     try:
#         async with httpx.AsyncClient(timeout=10.0) as client:
#             response = await client.get(f"{OLLAMA_URL}/api/tags")
            
#             if response.status_code != 200:
#                 return {"status": "error", "message": "Ollama not accessible"}
            
#             models = response.json()
#             available_models = [model["name"] for model in models.get("models", [])]
            
#             model_available = any(MODEL_NAME in model for model in available_models)
            
#             return {
#                 "status": "healthy" if model_available else "model_missing",
#                 "ollama_running": True,
#                 "model_name": MODEL_NAME,
#                 "model_available": model_available,
#                 "available_models": available_models
#             }
            
#     except Exception as e:
#         return {
#             "status": "error",
#             "message": f"Cannot connect to Ollama: {str(e)}",
#             "ollama_running": False
#         }

@app.post("/upload-csv", response_model=JobResponse)
async def upload_csv(
    file: UploadFile = File(...),
    system_prompt: str = Form(default=DEFAULT_SYSTEM_PROMPT),
    user_prompt: str = Form(default="Please analyze this CSV file and provide insights.")
):
    """
    Upload a CSV file and analyze it with Llama
    """
    try:
        # Validate file
        if file.content_type != "text/csv":
            raise HTTPException(
                status_code=400, 
                detail=f"File {file.filename} is not a CSV. Only CSV files are allowed"
            )
        
        # Read file content
        file_content = await file.read()
        
        # Validate file size
        if len(file_content) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"File {file.filename} exceeds {MAX_FILE_SIZE / (1024*1024)}MB limit"
            )
        
        # Generate job ID and create structure
        job_id = get_next_job_id()
        job_dir = create_job_structure(job_id)
        
        logger.info(f"Created job {job_id} for CSV file {file.filename}")
        
        # Save job metadata
        job_metadata = {
            "job_id": job_id,
            "created_at": datetime.now().isoformat(),
            "total_files": 1,
            "system_prompt": system_prompt,
            "user_prompt": user_prompt,
            "status": "processing",
            "files": [{"filename": file.filename, "status": "pending"}]
        }
        
        with open(job_dir / "job_metadata.json", "w", encoding="utf-8") as f:
            json.dump(job_metadata, f, indent=2, ensure_ascii=False)
        
        # Save the CSV file
        csv_file_path = job_dir / "uploads" / file.filename
        with open(csv_file_path, "wb") as buffer:
            buffer.write(file_content)
        
        logger.info(f"Saved CSV file to {csv_file_path}")
        
        # Process the CSV file and analyze
        await process_csv_file(job_id, job_dir, csv_file_path, system_prompt, user_prompt)
        
        return JobResponse(
            job_id=job_id,
            status="completed",
            message="CSV file processed and analyzed successfully.",
            files_processed=1,
            created_at=job_metadata["created_at"]
        )
        
    except Exception as e:
        logger.error(f"Error processing CSV file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error processing CSV file: {str(e)}")

async def process_csv_file(job_id: str, job_dir: Path, csv_file_path: Path, system_prompt: str, user_prompt: str):
    """Process and analyze the CSV file"""
    try:
        logger.info(f"Processing CSV file {csv_file_path} for job {job_id}")
        
        # Read the CSV file
        df = pd.read_csv(csv_file_path)
        
        # Convert the DataFrame to a string for analysis
        csv_content = df.to_csv(index=False, quoting=csv.QUOTE_NONNUMERIC)
        
        # Use OpenRouter to analyze the CSV content
        openrouter_response = await process_with_openrouter(
            csv_content,
            system_prompt,
            user_prompt,
            job_id,
            job_dir / "processed_results"
        )
        
        # Handle case where OpenRouter might fail
        if isinstance(openrouter_response, Exception):
            logger.error(f"OpenRouter processing failed for CSV: {str(openrouter_response)}")
            openrouter_response = {
                "error": str(openrouter_response),
                "response": f"Error processing CSV with OpenRouter: {str(openrouter_response)}"
            }
        
        # Save the analysis results
        results_file = job_dir / "processed_results" / f"{job_id}_csv_analysis.json"
        with open(results_file, "w", encoding="utf-8") as f:
            json.dump(openrouter_response, f, indent=2, ensure_ascii=False)
        
        logger.info(f"CSV analysis results saved to {results_file}")
        
    except Exception as e:
        logger.error(f"Error processing CSV file {csv_file_path}: {str(e)}")
        raise Exception(f"CSV processing failed for {csv_file_path}: {str(e)}")


@app.post("/upload-multiple-pdfs-with-ledger-xlsx")
async def upload_multiple_pdfs_with_ledger_xlsx(
    files: List[UploadFile] = File(...),
    ledger_file: UploadFile = File(...),
    system_prompt: str = Form(default=DEFAULT_SYSTEM_PROMPT),
    user_prompt: str = Form(default="Extract information from these documents and prepare it for Excel export.")
):
    """
    Upload multiple PDF files and an Excel ledger, process PDFs with OCR,
    analyze with DeepSeek, and update the Excel ledger.
    """
    try:
        # Validate ledger file type
        if ledger_file.content_type != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            raise HTTPException(
                status_code=400,
                detail=f"Invalid ledger file format. Please upload an Excel XLSX file."
            )

        # Read and validate the ledger XLSX
        ledger_content = await ledger_file.read()
        try:
            # Temporarily save the file to read it with pandas
            temp_xlsx_path = Path("temp_ledger.xlsx")
            with open(temp_xlsx_path, "wb") as f:
                f.write(ledger_content)
            
            # Read with pandas to validate format
            ledger_df = pd.read_excel(temp_xlsx_path)
            logger.info(f"Ledger XLSX loaded with {len(ledger_df)} rows and columns: {', '.join(ledger_df.columns)}")
            
            # Clean up temp file
            if temp_xlsx_path.exists():
                temp_xlsx_path.unlink()
                
        except Exception as e:
            # Clean up temp file if it exists
            if temp_xlsx_path.exists():
                temp_xlsx_path.unlink()
                
            raise HTTPException(
                status_code=400,
                detail=f"Invalid Excel ledger format: {str(e)}"
            )
        
        # Validate PDFs
        pdf_files = []
        for file in files:
            if file.content_type != "application/pdf":
                raise HTTPException(
                    status_code=400, 
                    detail=f"File {file.filename} is not a PDF. Only PDF files are allowed"
                )
            
            # Read file content
            file_content = await file.read()
            
            # Validate file size
            if len(file_content) > MAX_FILE_SIZE:
                raise HTTPException(
                    status_code=400,
                    detail=f"File {file.filename} exceeds {MAX_FILE_SIZE / (1024*1024)}MB limit"
                )
            
            pdf_files.append({
                "filename": file.filename,
                "content": file_content
            })
        
        # Generate job ID and create structure
        job_id = get_next_job_id()
        job_dir = create_job_structure(job_id)
        
        # Save ledger XLSX
        ledger_path = job_dir / "ledger.xlsx"
        with open(ledger_path, "wb") as f:
            f.write(ledger_content)
        
        logger.info(f"Created job {job_id} for {len(pdf_files)} files with Excel ledger integration")
        
        # Save job metadata
        job_metadata = {
            "job_id": job_id,
            "created_at": datetime.now().isoformat(),
            "total_files": len(pdf_files),
            "system_prompt": system_prompt,
            "user_prompt": user_prompt,
            "status": "processing",
            "ledger_file": ledger_file.filename,
            "ledger_format": "xlsx",
            "files": [{"filename": f["filename"], "status": "pending"} for f in pdf_files]
        }
        
        with open(job_dir / "job_metadata.json", "w", encoding="utf-8") as f:
            json.dump(job_metadata, f, indent=2, ensure_ascii=False)
        
        # Process files and update ledger asynchronously
        asyncio.create_task(process_job_files_with_ledger_xlsx(
            job_id, job_dir, pdf_files, ledger_path, system_prompt, user_prompt
        ))
        
        return {
            "job_id": job_id,
            "status": "processing",
            "message": f"Job created successfully. Processing {len(pdf_files)} files with Excel ledger integration.",
            "files_processed": 0,
            "created_at": job_metadata["created_at"]
        }
        
    except Exception as e:
        logger.error(f"Error creating job with Excel ledger: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error creating job with Excel ledger: {str(e)}")

async def process_job_files_with_ledger_xlsx(job_id: str, job_dir: Path, pdf_files: List, ledger_path: Path, system_prompt: str, user_prompt: str):
    """Process files and update the Excel ledger with extracted data"""
    try:
        # Create a backup of the original ledger
        backup_path = ledger_path.parent / f"ledger_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        shutil.copy2(ledger_path, backup_path)
        logger.info(f"Created Excel ledger backup at {backup_path}")
        
        # Process files and get OCR results (using same code as CSV version)
        logger.info(f"Starting processing for job {job_id} with Excel ledger integration")
        
        # Update job status
        update_job_status(job_dir, "processing")
        
        all_ocr_results = []
        processed_files = []
        failed_files = []
        
        # Process each PDF file (existing OCR processing code)
        for i, pdf_file in enumerate(pdf_files, 1):
            try:
                logger.info(f"Processing file {i}/{len(pdf_files)}: {pdf_file['filename']}")
                
                # Generate unique filename and save
                file_extension = Path(pdf_file["filename"]).suffix
                unique_filename = f"file_{i}_{uuid.uuid4()}{file_extension}"
                file_path = job_dir / "uploads" / unique_filename
                
                with open(file_path, "wb") as buffer:
                    buffer.write(pdf_file["content"])
                
                logger.info(f"Saved file to {file_path}")
                
                # Process with enhanced OCR - reuse existing code
                try:
                    logger.info(f"Starting enhanced OCR for {file_path}")
                    ocr_result = await process_pdf_with_enhanced_ocr(
                        file_path, 
                        f"{job_id}_file_{i}", 
                        job_dir / "ocr_results",
                        original_filename=pdf_file["filename"]
                    )
                    
                    if not ocr_result or not ocr_result.get("full_text", "").strip():
                        logger.warning(f"OCR produced empty result for {pdf_file['filename']}")
                        ocr_result["full_text"] = f"[OCR EXTRACTION FAILED FOR {pdf_file['filename']}]"
                    
                    all_ocr_results.append(ocr_result)
                    processed_files.append({
                        "original_filename": pdf_file["filename"],
                        "stored_filename": unique_filename,
                        "file_number": i,
                        "ocr_result": ocr_result,
                        "status": "success",
                        "tables_extracted": len(ocr_result.get("tables", [])) 
                    })
                    
                    logger.info(f"Completed enhanced OCR for file {i}/{len(pdf_files)} with {len(ocr_result.get('tables', []))} tables extracted")
                except Exception as ocr_error:
                    logger.error(f"OCR processing failed for {pdf_file['filename']}: {str(ocr_error)}")
                    failed_files.append({
                        "original_filename": pdf_file["filename"],
                        "stored_filename": unique_filename,
                        "file_number": i,
                        "error": str(ocr_error),
                        "status": "failed"
                    })
                    continue
                
            except Exception as e:
                logger.error(f"Error processing file {pdf_file['filename']}: {str(e)}")
                failed_files.append({
                    "original_filename": pdf_file["filename"],
                    "file_number": i,
                    "error": str(e),
                    "status": "failed"
                })
                continue
        
        # Check if we have any successful OCR results
        if not processed_files:
            error_msg = "All files failed OCR processing"
            logger.error(error_msg)
            update_job_status(job_dir, "failed", error_msg)
            return
        
        # Combine all OCR results for collective analysis
        combined_context = prepare_combined_context(processed_files)
        
        # Process with OpenRouter (same as CSV version)
        openrouter_response = await process_with_openrouter(
            combined_context,
            system_prompt,
            user_prompt,
            job_id,
            job_dir / "processed_results"
        )
        
        # Handle case where OpenRouter might fail
        if isinstance(openrouter_response, Exception):
            logger.error(f"OpenRouter processing failed: {str(openrouter_response)}")
            openrouter_response = {
                "error": str(openrouter_response),
                "response": f"Error processing with OpenRouter: {str(openrouter_response)}"
            }
        
        # Extract data using the same function as CSV
        extracted_data = extract_csv_data_from_response(openrouter_response.get('response', ''))
        
        # Save extracted data for debugging
        with open(job_dir / "extracted_csv_data.json", "w", encoding="utf-8") as f:
            json.dump(extracted_data, f, indent=2, ensure_ascii=False)
        
        # Update the Excel ledger with the extracted data using the new function
        updated_ledger_path = job_dir / "updated_ledger.xlsx"
        
        try:
            # Update the original ledger file
            updated_workbook = update_ledger_xlsx(ledger_path, extracted_data)
            
            # Save a copy in the job directory
            updated_workbook.save(updated_ledger_path)
            
            logger.info(f"Updated Excel ledger saved to {updated_ledger_path}")
            
            # Create job summary
            job_summary = {
                "job_id": job_id,
                "status": "completed",
                "completed_at": datetime.now().isoformat(),
                "total_files": len(pdf_files),
                "successfully_processed": len(processed_files),
                "failed_files": len(pdf_files) - len(processed_files),
                "processed_files": processed_files,
                "llm_analysis": openrouter_response,
                "extracted_csv_data": extracted_data,
                "ledger_update": {
                    "invoice_rows_added": len(extracted_data.get("csv_data", {}).get("invoice_rows", [])),
                    "delivery_note_rows_added": len(extracted_data.get("csv_data", {}).get("delivery_note_rows", [])),
                    "anomaly_rows_added": len(extracted_data.get("csv_data", {}).get("anomaly_rows", [])),
                    "updated_ledger_path": str(updated_ledger_path),
                    "format": "xlsx"
                }
            }
            
            # Save the job results
            with open(job_dir / "job_results.json", "w", encoding="utf-8") as f:
                json.dump(job_summary, f, default=str, indent=2)
                
            update_job_status(job_dir, "completed")
            logger.info(f"Job {job_id} completed successfully with Excel data integration")
            return job_summary
            
        except Exception as e:
            logger.error(f"Error updating Excel ledger: {str(e)}")
            update_job_status(job_dir, "failed", f"Excel ledger update failed: {str(e)}")
            raise
            
    except Exception as e:
        logger.error(f"Error processing job with Excel ledger: {str(e)}")
        update_job_status(job_dir, "failed", str(e))
        raise

def update_ledger_xlsx(ledger_path: Path, extracted_data: dict) -> openpyxl.Workbook:
    """Update the Excel ledger with extracted data from documents in structured format"""
    try:
        # Check if openpyxl is available
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel processing but not installed")
        
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Read the Excel file
        workbook = load_workbook(filename=str(ledger_path))
        
        # Get the active sheet
        sheet = workbook.active
        
        # Current date for posting date
        current_date = datetime.now().strftime("%m/%d/%Y")
        
        # Define styles
        try:
            anomaly_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            anomaly_font = Font(color="9C6500")
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        except Exception as e:
            logger.warning(f"Could not create styles: {e}")
            # Create dummy styles that won't be used but prevent errors
            anomaly_fill = None
            anomaly_font = None
            border = None
        
        # Group related data for integrated rows
        invoice_rows = extracted_data.get("csv_data", {}).get("invoice_rows", [])
        delivery_rows = extracted_data.get("csv_data", {}).get("delivery_note_rows", [])
        anomaly_rows = extracted_data.get("csv_data", {}).get("anomaly_rows", [])
        
        # Organize delivery notes by invoice number for easier matching
        delivery_by_invoice = {}
        for delivery_row in delivery_rows:
            delivery_parts = [part.strip() for part in delivery_row.split(',')]
            if len(delivery_parts) >= 5 and delivery_parts[2]:  # Check if has invoice number
                delivery_by_invoice[delivery_parts[2]] = delivery_parts
        
        # Process each invoice and create integrated rows
        next_row = sheet.max_row + 1
        rows_added = 0
        
        # Process invoices and match with delivery notes
        for invoice_row in invoice_rows:
            invoice_parts = [part.strip() for part in invoice_row.split(',')]
            
            if len(invoice_parts) >= 6:
                invoice_id = invoice_parts[1]
                
                # Create a new empty row
                for col_idx in range(1, sheet.max_column + 1):
                    sheet.cell(row=next_row, column=col_idx, value="")
                
                # General Ledger Data (columns 1-11)
                sheet.cell(row=next_row, column=1, value=f"Sale of goods - Invoice {invoice_id}")  # Transaction Description
                sheet.cell(row=next_row, column=2, value=invoice_parts[0])  # Transaction Date (invoice date)
                sheet.cell(row=next_row, column=3, value=current_date)  # Posting Date
                sheet.cell(row=next_row, column=4, value=invoice_parts[2])  # Customer Name
                sheet.cell(row=next_row, column=5, value=invoice_id)  # Invoice ID
                sheet.cell(row=next_row, column=6, value=invoice_parts[0])  # Invoice Date
                sheet.cell(row=next_row, column=7, value=invoice_parts[3])  # Amount (No VAT)
                sheet.cell(row=next_row, column=8, value=invoice_parts[4])  # VAT Amount
                sheet.cell(row=next_row, column=9, value=invoice_parts[5])  # Total Amount
                sheet.cell(row=next_row, column=10, value="system")  # Posted By
                
                # Invoice data (columns 13-19)
                sheet.cell(row=next_row, column=13, value=invoice_parts[0])  # Invoice date
                sheet.cell(row=next_row, column=14, value=invoice_parts[1])  # Invoice ID
                sheet.cell(row=next_row, column=15, value=invoice_parts[2])  # Customer name
                sheet.cell(row=next_row, column=16, value=invoice_parts[3])  # Invoice amount (No VAT)
                sheet.cell(row=next_row, column=17, value=invoice_parts[4])  # VAT
                sheet.cell(row=next_row, column=19, value=invoice_parts[5])  # Total Amount
                
                # Check if there's a matching delivery note
                matched_delivery = delivery_by_invoice.get(invoice_id)
                if matched_delivery:
                    # Add delivery note data to the same row (columns 20-24)
                    sheet.cell(row=next_row, column=20, value=matched_delivery[0])  # Delivery note date
                    sheet.cell(row=next_row, column=21, value=matched_delivery[1])  # Delivery note number
                    sheet.cell(row=next_row, column=22, value=matched_delivery[2])  # Invoice number
                    sheet.cell(row=next_row, column=23, value=matched_delivery[3])  # Invoice date
                    sheet.cell(row=next_row, column=24, value=matched_delivery[4])  # Customer name
                
                # Check for anomalies related to this invoice
                related_anomalies = []
                for anomaly in anomaly_rows:
                    if invoice_id in anomaly:
                        related_anomalies.append(anomaly)
                
                if related_anomalies:
                    combined_anomalies = "; ".join(related_anomalies)
                    sheet.cell(row=next_row, column=11, value=combined_anomalies)  # Anomaly Type
                    
                    # Apply styling if available
                    if anomaly_fill and anomaly_font:
                        try:
                            sheet.cell(row=next_row, column=11).fill = anomaly_fill
                            sheet.cell(row=next_row, column=11).font = anomaly_font
                        except Exception as style_error:
                            logger.warning(f"Could not apply anomaly style: {style_error}")
                
                # Apply borders if available
                if border:
                    for col_idx in range(1, sheet.max_column + 1):
                        try:
                            sheet.cell(row=next_row, column=col_idx).border = border
                        except Exception:
                            pass
                
                next_row += 1
                rows_added += 1
        
        # Handle any unmatched delivery notes
        matched_invoice_ids = [inv_parts[1] for inv_parts in [inv.split(',') for inv in invoice_rows] if len(inv_parts) >= 2]
        for delivery_row in delivery_rows:
            delivery_parts = [part.strip() for part in delivery_row.split(',')]
            
            if len(delivery_parts) >= 5:
                invoice_num = delivery_parts[2]
                
                # Only process if this delivery note hasn't been matched to an invoice
                if not invoice_num or invoice_num not in matched_invoice_ids:
                    # Create a new empty row
                    for col_idx in range(1, sheet.max_column + 1):
                        sheet.cell(row=next_row, column=col_idx, value="")
                    
                    # General Ledger Data
                    sheet.cell(row=next_row, column=1, value=f"Delivery Note - {delivery_parts[1]}")  # Transaction Description
                    sheet.cell(row=next_row, column=2, value=delivery_parts[0])  # Transaction Date
                    sheet.cell(row=next_row, column=3, value=current_date)  # Posting Date
                    if delivery_parts[4]:
                        sheet.cell(row=next_row, column=4, value=delivery_parts[4])  # Customer Name
                    if delivery_parts[2]:
                        sheet.cell(row=next_row, column=5, value=delivery_parts[2])  # Invoice ID reference
                    sheet.cell(row=next_row, column=10, value="system")  # Posted By
                    
                    # Delivery note data (columns 20-24)
                    sheet.cell(row=next_row, column=20, value=delivery_parts[0])  # Delivery note date
                    sheet.cell(row=next_row, column=21, value=delivery_parts[1])  # Delivery note number
                    sheet.cell(row=next_row, column=22, value=delivery_parts[2])  # Invoice number
                    sheet.cell(row=next_row, column=23, value=delivery_parts[3])  # Invoice date
                    sheet.cell(row=next_row, column=24, value=delivery_parts[4])  # Customer name
                    
                    # Apply borders if available
                    if border:
                        for col_idx in range(1, sheet.max_column + 1):
                            try:
                                sheet.cell(row=next_row, column=col_idx).border = border
                            except Exception:
                                pass
                    
                    next_row += 1
                    rows_added += 1
        
        # Handle orphaned anomalies (those not tied to specific invoice/delivery)
        matched_anomalies = []
        for invoice_row in invoice_rows:
            invoice_parts = [part.strip() for part in invoice_row.split(',')]
            if len(invoice_parts) >= 2:
                invoice_id = invoice_parts[1]
                for anomaly in anomaly_rows:
                    if invoice_id in anomaly:
                        matched_anomalies.append(anomaly)
        
        unmatched_anomalies = [a for a in anomaly_rows if a not in matched_anomalies]
        if unmatched_anomalies:
            # Create a new empty row
            for col_idx in range(1, sheet.max_column + 1):
                sheet.cell(row=next_row, column=col_idx, value="")
            
            # Add combined anomalies
            combined_anomalies = "; ".join(unmatched_anomalies)
            sheet.cell(row=next_row, column=1, value="Anomaly Report")  # Transaction Description
            sheet.cell(row=next_row, column=2, value=current_date)  # Transaction Date
            sheet.cell(row=next_row, column=3, value=current_date)  # Posting Date
            sheet.cell(row=next_row, column=10, value="system")  # Posted By
            sheet.cell(row=next_row, column=11, value=combined_anomalies)  # Anomaly Type
            
            # Apply styling if available
            if anomaly_fill and anomaly_font:
                try:
                    sheet.cell(row=next_row, column=11).fill = anomaly_fill
                    sheet.cell(row=next_row, column=11).font = anomaly_font
                except Exception as style_error:
                    logger.warning(f"Could not apply anomaly style: {style_error}")
            
            # Apply borders if available
            if border:
                for col_idx in range(1, sheet.max_column + 1):
                    try:
                        sheet.cell(row=next_row, column=col_idx).border = border
                    except Exception:
                        pass
                    
            next_row += 1
            rows_added += 1
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            adjusted_width = (max_length + 2) if max_length > 0 else 10
            sheet.column_dimensions[column_letter].width = min(adjusted_width, 30)  # Cap at 30
        
        logger.info(f"Added {rows_added} integrated rows to Excel ledger")
        
        # Return the updated workbook
        return workbook
        
    except Exception as e:
        logger.error(f"Error updating Excel ledger: {str(e)}")
        raise


@app.get("/ledger/csv")
def download_global_ledger_csv():
    """Download the single-table global ledger CSV."""
    if not GLOBAL_LEDGER_CSV.exists():
        raise HTTPException(status_code=404, detail="Global ledger CSV not found")
    return FileResponse(
        path=str(GLOBAL_LEDGER_CSV),
        media_type="text/csv",
        filename="global_ledger.csv",
    )

# Optional: alternate path your frontend might be calling
@app.get("/download/global-ledger")
def download_global_ledger_csv_alias():
    if not GLOBAL_LEDGER_CSV.exists():
        raise HTTPException(status_code=404, detail="Global ledger CSV not found")
    return FileResponse(
        path=str(GLOBAL_LEDGER_CSV),
        media_type="text/csv",
        filename="global_ledger.csv",
    )


# Function to convert date string to Excel serial date
def date_to_serial(date_str):
    if not date_str:
        return None
    base_date = datetime(1899, 12, 30)
    dt = datetime.strptime(date_str, '%Y-%m-%d')
    return (dt - base_date).days

# Step 1: Match ledger and extracted invoice data
def match_ledger_invoice(ledger_row, extracted_invoice):
    matches = {}
    inv_date_serial = date_to_serial(extracted_invoice['Invoice date'])
    matches['Invoice Date'] = ledger_row['Transaction Date'] == inv_date_serial if inv_date_serial is not None else False
    matches['Customer Name'] = ledger_row['Customer Name'] == extracted_invoice['Customer name']
    matches['Amount (No VAT)'] = abs(float(ledger_row['Amount (No VAT)']) - extracted_invoice['Invoice amount (No VAT)']) < 0.01
    matches['VAT Amount'] = abs(float(ledger_row['VAT Amount']) - extracted_invoice['VAT']) < 0.01
    matches['Total Amount'] = abs(float(ledger_row['Total Amount']) - extracted_invoice['Total Amount']) < 0.01
    
    overall_status = 'Match' if all(matches.values()) else 'Not Matched'
    return matches, overall_status

# Step 2: Match extracted invoice and delivery note
def match_invoice_delivery(extracted_invoice, extracted_delivery_note):
    matches = {}
    matches['Invoice ID'] = extracted_invoice['Invoice ID'] == extracted_delivery_note['Invoice number']
    matches['Invoice Date'] = extracted_invoice['Invoice date'] == extracted_delivery_note['Invoice date']
    matches['Customer Name'] = extracted_invoice['Customer name'] == extracted_delivery_note['Customer name']
    
    overall_status = 'Match' if all(matches.values()) else 'Not Matched'
    return matches, overall_status

# Function to update ledger with extracted data from a job
def update_ledger_with_extracted_data(extracted_data):
    if not LEDGER_FILE.exists():
        raise HTTPException(status_code=404, detail="Ledger file not found")
    
    # Load the Excel file with pandas for data manipulation
    with pd.ExcelFile(LEDGER_FILE) as xls:
        df_teting = pd.read_excel(xls, sheet_name='Teting Sheet', header=1)
    
    # Use openpyxl for precise cell updates
    wb = openpyxl.load_workbook(LEDGER_FILE)
    ws = wb['Teting Sheet']
    
    # Process each extracted invoice
    updated = False
    for inv_row in extracted_data['csv_data']['invoice_rows']:
        parts = [p.strip() for p in inv_row.split(',')]
        if len(parts) < 6:
            continue
        extracted_invoice = {
            'Invoice date': parts[0],
            'Invoice ID': parts[1],
            'Customer name': parts[2],
            'Invoice amount (No VAT)': float(parts[3]),
            'VAT': float(parts[4]),
            'Total Amount': float(parts[5])
        }
        
        # Find matching delivery note by Invoice number
        extracted_delivery_note = None
        for dn_row in extracted_data['csv_data']['delivery_note_rows']:
            dn_parts = [p.strip() for p in dn_row.split(',')]
            if len(dn_parts) >= 5 and dn_parts[2] == extracted_invoice['Invoice ID']:
                extracted_delivery_note = {
                    'Delivery note date': dn_parts[0],
                    'Delivery note number': dn_parts[1],
                    'Invoice number': dn_parts[2],
                    'Invoice date': dn_parts[3],
                    'Customer name': dn_parts[4]
                }
                break
        
        if not extracted_delivery_note:
            logger.warning(f"No matching delivery note for invoice {extracted_invoice['Invoice ID']}")
            continue
        
        # Find row in Teting Sheet where Invoice ID matches
        match_row = df_teting[df_teting['Invoice ID'] == extracted_invoice['Invoice ID']].index
        if len(match_row) == 0:
            logger.warning(f"No matching row found for Invoice ID: {extracted_invoice['Invoice ID']}")
            continue
        
        match_row_idx = match_row[0] + 2  # Adjust for header (1-based, +1 for openpyxl)
        ledger_row = df_teting.iloc[match_row[0]].to_dict()
        
        # Perform matches and log
        ledger_matches, ledger_status = match_ledger_invoice(ledger_row, extracted_invoice)
        inv_dn_matches, inv_dn_status = match_invoice_delivery(extracted_invoice, extracted_delivery_note)
        logger.info(f"Ledger match for {extracted_invoice['Invoice ID']}: {ledger_status}")
        logger.info(f"Invoice-Delivery match: {inv_dn_status}")
        
        # If mismatch, append to Anomaly Type (column 11)
        if ledger_status == 'Not Matched' or inv_dn_status == 'Not Matched':
            anomaly_cell = ws.cell(row=match_row_idx + 1, column=11)
            current_anomaly = anomaly_cell.value or ""
            new_anomaly = f"Ledger Mismatch: {ledger_status}; Invoice-Delivery Mismatch: {inv_dn_status}"
            anomaly_cell.value = f"{current_anomaly}; {new_anomaly}" if current_anomaly else new_anomaly
            
            # Optional styling
            anomaly_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            anomaly_font = Font(color="9C6500")
            anomaly_cell.fill = anomaly_fill
            anomaly_cell.font = anomaly_font
        
        # Convert dates to serial numbers
        inv_date_serial = date_to_serial(extracted_invoice['Invoice date'])
        dn_date_serial = date_to_serial(extracted_delivery_note['Invoice date'])
        dn_note_date_serial = date_to_serial(extracted_delivery_note['Delivery note date'])
        
        # Insert invoice data (columns M to R: 13 to 18)
        invoice_data = [
            inv_date_serial,
            extracted_invoice['Invoice ID'],
            extracted_invoice['Customer name'],
            extracted_invoice['Invoice amount (No VAT)'],
            extracted_invoice['VAT'],
            extracted_invoice['Total Amount']
        ]
        for col, value in enumerate(invoice_data, start=13):
            ws.cell(row=match_row_idx + 1, column=col, value=value)
        
        # Insert delivery data (columns T to X: 20 to 24)
        delivery_data = [
            dn_note_date_serial,
            extracted_delivery_note['Delivery note number'],
            extracted_delivery_note['Invoice number'],
            dn_date_serial,
            extracted_delivery_note['Customer name']
        ]
        for col, value in enumerate(delivery_data, start=20):
            ws.cell(row=match_row_idx + 1, column=col, value=value)
        
        # Apply borders to the row
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
        for col in range(1, ws.max_column + 1):
            ws.cell(row=match_row_idx + 1, column=col).border = border
        
        updated = True
    
    if updated:
        wb.save(LEDGER_FILE)
        logger.info(f"Ledger updated successfully for job")
    else:
        logger.info("No updates made to ledger - no matching entries found")



# Add download endpoint
@app.get("/ledger/download")
def download_ledger():
    if not LEDGER_FILE.exists():
        raise HTTPException(status_code=404, detail="Ledger file not found")
    return FileResponse(
        path=str(LEDGER_FILE),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="ledger.xlsx"
    )